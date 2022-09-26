from faker import Faker
from openpyxl import Workbook

"""все методы вроде не описаны в документации, можно юзать dir
https://faker.readthedocs.io/en/master/index.html# 
"""
# choose localization - a lot of them in documentation
fake_data = Faker(["fi_FI", "en"])
# fake_data = Faker()
# print all methods
# print(dir(fake_data))
wb = Workbook()
ws = wb.active

# Запись в эксель файл сгенерирорванных данных
for i in range(1, 11):
    for j in range(1, 4):
        ws.cell(row=i, column=1).value = fake_data.first_name()
        ws.cell(row=i, column=2).value = fake_data.last_name()
        ws.cell(row=i, column=3).value = fake_data.name()
        ws.cell(row=i, column=4).value = fake_data.email()
wb.save("testdata.xlsx")



# Работа с одинаковыми значениями через seed_instance, ещё вроде можно через seed
fake1 = Faker()
fake1.seed_instance(0)
fake2 = Faker()
fake2.seed_instance(0)
print(fake1.name())
print(fake2.name())


